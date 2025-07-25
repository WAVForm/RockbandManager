import requests
import openpyxl
import logging
import xlsxwriter

class XLSXManager:
    # TODO database manager
    # DB_FILE = "songs.db"
    OUTPUT_FILE = "songs.xlsx"

    def __init__(self, DL_URL):
        self.logger = logging.getLogger('XLSXManager')
        self.DL_URL = DL_URL

    def download_google_sheet_xlsx(self, max_attempts=10):
        for attempt in range(max_attempts):
            try:
                self.logger.info(f"[XLSXManager] Downloading Google Sheet from {self.DL_URL} ...")
                response = requests.get(self.DL_URL)
                if response.status_code == 200:
                    with open(self.OUTPUT_FILE, "wb") as f:
                        f.write(response.content)
                    self.logger.info(f"[XLSXManager] Downloaded sheet to {self.OUTPUT_FILE}")
                else:
                    raise requests.exceptions.ConnectionError("Did not get status 200")
            except Exception as e:
                self.logger.error(f"[XLSXManager] Error downloading Google Sheet, attempt {attempt}: {e}")

    def get_all_customs_file_ids_from_worksheet(self):
        wb = openpyxl.load_workbook(self.OUTPUT_FILE, data_only=True)
        ws = wb['customs']
        
        file_ids = []
        for row in ws.iter_rows(min_row=2):
            file_id_cell = row[4].value  # Column E (File ID)
            file_ids.append(file_id_cell)

        return file_ids

    def get_file_ids_not_in_xlsx(self):
        try:
            # TODO database manager
            # conn = sqlite3.connect(self.DB_FILE)
            # cursor = conn.cursor()
            # cursor.execute("SELECT file_id FROM customs")
            # db_file_ids = {row[0] for row in cursor.fetchall()}
            # conn.close()

            new_file_ids = [file_id for file_id in all_customs_file_ids if file_id not in self.get_all_customs_file_ids_from_worksheet()]
            self.logger.info(f"[XLSXManager] Found {len(new_file_ids)} new file IDs.")
            return new_file_ids
        except Exception as e:
            self.logger.error(f"[XLSXManager] Error fetching new file IDs: {e}")
            return []

    @staticmethod
    def is_full_band(song):
        return all(song[k] not in (None, -1) for k in ["diff_drums", "diff_guitar", "diff_bass", "diff_vocals"])

    def update_wanted_from_worksheets(self):
        try:
            wb = load_workbook(self.OUTPUT_FILE, data_only=True)
            customs_sheet = wb["customs"]
            songs_sheet = wb["songs"]

            # TODO database manager
            # conn = sqlite3.connect(self.DB_FILE)
            # cursor = conn.cursor()

            for row in customs_sheet.iter_rows(min_row=2):
                file_id = row[4].value  # Column E (File ID)
                wanted = row[0].value  # Column A (Wanted?)

                # TODO database manager
                # if file_id and wanted is not None:
                #     cursor.execute("UPDATE customs SET wanted = ? WHERE file_id = ?", (bool(wanted), file_id))

            for row in songs_sheet.iter_rows(min_row=2):
                title = row[1].value  # Column B (Song Name)
                artist = row[2].value  # Column C (Song Artist)
                wanted = row[0].value  # Column A (Wanted?)
                
                # TODO database manager
                # if title and artist and wanted is not None:
                #     cursor.execute("UPDATE songs SET wanted = ? WHERE title = ? AND artist = ?", (bool(wanted), title, artist))

            # conn.commit()
            # conn.close()

            self.logger.info("[XLSXManager] Updated 'wanted' field in the database from the sheets.")
        except Exception as e:
            self.logger.error(f"[XLSXManager] Error updating 'wanted' field in the database: {e}")

    def export_customs(self):
        try:
            # TODO database manager
            # conn = sqlite3.connect(self.DB_FILE)
            # cursor = conn.cursor()
            # cursor.execute("SELECT wanted, file_id, title, artist, diff_drums, diff_guitar, diff_bass, diff_vocals FROM customs")
            # rows = cursor.fetchall()
            # conn.close()

            workbook = xlsxwriter.Workbook(self.OUTPUT_FILE)
            worksheet = workbook.add_worksheet("customs")

            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#D9E1F2',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })

            text_format = workbook.add_format({
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })

            warn_format = workbook.add_format({
                'border': 1,
                'align': 'center',
                'valign': 'vcenter',
                'font_size': 20,
                'bold': True,             
                'font_color': 'red'
            })

            headers = ["Wanted?", "Song Name", "Song Artist", "Full Band?", "File ID"]
            worksheet.write_row(0, 0, headers, header_format)

            # Set column widths (adjust as needed)
            worksheet.set_column('A:A', 10)
            worksheet.set_column('B:C', 40)
            worksheet.set_column('D:D', 12)
            worksheet.set_column('E:E', 40)

            worksheet.set_row(0,48)

            max_row = len(rows) + 1

            green_fill = workbook.add_format({'bg_color': '#C6EFCE'})
            red_fill = workbook.add_format({'bg_color': '#FFC7CE'})

            worksheet.conditional_format(f"A2:E{max_row}", {
                'type': 'formula',
                'criteria': '$A2=TRUE',
                'format': green_fill
            })
            worksheet.conditional_format(f"A2:E{max_row}", {
                'type': 'formula',
                'criteria': '$A2=FALSE',
                'format': red_fill
            })

            for i, row in enumerate(rows, start=1):
                wanted, file_id, title, artist, d1, d2, d3, d4 = row
                full_band = all(x not in (None, -1) for x in [d1, d2, d3, d4])

                worksheet.set_row(i,48)
                worksheet.insert_checkbox(i, 0, wanted)
                worksheet.write(i, 1, title, text_format)
                worksheet.write(i, 2, artist, text_format)
                worksheet.write(i, 3, str(full_band), text_format if full_band else warn_format)
                worksheet.write(i, 4, file_id, text_format)

            workbook.close()
            self.logger.info(f"[XLSXManager] Exported customs worksheet to {self.OUTPUT_FILE}.")
        except Exception as e:
            self.logger.info(f"[XLSXManager] Failed exporting customs worksheet: {e}")

    def export_songs(self):
        try:
            # TODO database manager
            # conn = sqlite3.connect(self.DB_FILE)
            # cursor = conn.cursor()
            # cursor.execute("SELECT wanted, title, artist FROM songs")
            # rows = cursor.fetchall()
            # conn.close()
            workbook = xlsxwriter.Workbook(self.OUTPUT_FILE)
            worksheet = workbook.add_worksheet("songs")

            header_format = workbook.add_format({
                'bold': True,
                'bg_color': '#D9E1F2',
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })

            text_format = workbook.add_format({
                'border': 1,
                'align': 'center',
                'valign': 'vcenter'
            })

            headers = ["Wanted?", "Song Name", "Song Artist"]
            worksheet.write_row(0, 0, headers, header_format)

            worksheet.set_column('A:A', 10)
            worksheet.set_column('B:C', 40)

            worksheet.set_row(0,48)

            max_row = len(rows) + 1

            green_fill = workbook.add_format({'bg_color': '#C6EFCE'})
            red_fill = workbook.add_format({'bg_color': '#FFC7CE'})

            worksheet.conditional_format(f"A2:E{max_row}", {
                'type': 'formula',
                'criteria': '$A2=TRUE',
                'format': green_fill
            })
            worksheet.conditional_format(f"A2:E{max_row}", {
                'type': 'formula',
                'criteria': '$A2=FALSE',
                'format': red_fill
            })

            for i, row in enumerate(rows, start=1):
                wanted, title, artist = row

                worksheet.set_row(i,48)
                worksheet.insert_checkbox(i, 0, wanted)
                worksheet.write(i, 1, title, text_format)
                worksheet.write(i, 2, artist, text_format)

            workbook.close()
            self.logger.info(f"[XLSXManager] Exported songs worksheet to {self.OUTPUT_FILE}.")
        except Exception as e:
            self.logger.info(f"[XLSXManager] Failed exporting songs worksheet: {e}")